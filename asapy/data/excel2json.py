import openpyxl
import json
import MeCab 
import CaboCha
import re

book = openpyxl.load_workbook('pth20210305.xlsx')
sheets = book['pth20210305-sjis']
path_w = 'json_test.json'

def remove_part(surface,part):
    cp = CaboCha.Parser()
    tree = cp.parse(surface)
    pos = re.split('[*]', tree.token(0).feature)
    pos = pos[0][:-1]
    try:
        spliter = re.split('[・?]',part)
    except TypeError:
        print("TypeERROR",surface,part, pos)
        return surface, part , pos
    for split in spliter:
        if surface != surface.rstrip(split):
            surface = surface.rstrip(split)
            part = split
    #print(surface,part, pos)
    return surface , part , pos 


for i in range(2,12):
    H_col = sheets.cell(row=i, column=8).value #部署が
    G_col = sheets.cell(row=i, column=7).value #が（助詞）
    surface, part , pos = remove_part(H_col,G_col)
    semrole = sheets.cell(row=i, column=6).value
	# name = sheets.cell(row=i, column=2).value
	# status = sheets.cell(row=i, column=3).value   
    cases = {"cases":[{"noun" : surface,"semrole":semrole,"part":part}]}

    with open(path_w, mode='a') as f:
        f.write(json.dumps(cases, sort_keys=False, ensure_ascii=False,indent=4))